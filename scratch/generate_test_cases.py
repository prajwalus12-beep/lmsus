import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_manual_test_cases_excel():
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Styles
    font_family = "Segoe UI"
    
    # Fonts
    title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    section_font = Font(name=font_family, size=12, bold=True, color="1F2937")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    body_font = Font(name=font_family, size=10, bold=False, color="374151")
    id_font = Font(name=font_family, size=10, bold=True, color="1E3A8A")
    meta_label_font = Font(name=font_family, size=10, bold=True, color="1F2937")
    meta_val_font = Font(name=font_family, size=10, bold=False, color="4B5563")
    instruction_font = Font(name=font_family, size=10, italic=True, color="4B5563")

    # Fills
    title_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Deep Dark Blue
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid") # Royal Blue
    sub_header_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid") # Very Light Blue
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") # Off-White
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Status Fills (for future use or templates)
    pass_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # Light Green
    fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Light Red
    pending_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid") # Light Amber

    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_top_center = Alignment(horizontal="center", vertical="top", wrap_text=True)

    # Borders
    thin_border_side = Side(border_style="thin", color="D1D5DB")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    thick_bottom_side = Side(border_style="medium", color="1E3A8A")
    header_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thick_bottom_side)

    # Standard columns
    cols_headers = ["ID", "Description", "Step", "Expected Result", "Actual Result", "Status"]
    col_widths = {
        "A": 15,  # ID
        "B": 35,  # Description
        "C": 45,  # Step
        "D": 45,  # Expected Result
        "E": 20,  # Actual Result
        "F": 12   # Status
    }

    def format_sheet_common(ws):
        # Set column widths
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
        # Set gridlines visible
        ws.views.sheetView[0].showGridLines = True

    def write_header(ws, title_text):
        # Merge cells for Title
        ws.merge_cells("A1:F1")
        ws["A1"] = title_text
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 40

        # Subtitle or description row
        ws.merge_cells("A2:F2")
        ws["A2"] = "Manual Test Case Execution Sheet - Leave Management System (LMS)"
        ws["A2"].font = instruction_font
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 20

        # Empty row
        ws.append([])
        ws.row_dimensions[3].height = 10

        # Column Headers
        ws.append(cols_headers)
        header_row = 4
        ws.row_dimensions[header_row].height = 28
        for col_idx in range(1, 7):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = header_border

    def write_test_cases(ws, test_cases_list):
        current_row = 5
        for idx, tc in enumerate(test_cases_list):
            ws.append([
                tc["id"],
                tc["description"],
                tc["step"],
                tc["expected"],
                "", # Actual Result is blank for execution
                "Pending" # Default status
            ])
            
            # Formatting the row
            fill = zebra_fill if idx % 2 == 0 else white_fill
            ws.row_dimensions[current_row].height = 80 # default height for multi-line cells
            
            # Cell A: ID
            c_id = ws.cell(row=current_row, column=1)
            c_id.font = id_font
            c_id.alignment = align_top_center
            c_id.fill = fill
            c_id.border = thin_border
            
            # Cell B: Description
            c_desc = ws.cell(row=current_row, column=2)
            c_desc.font = body_font
            c_desc.alignment = align_top_left
            c_desc.fill = fill
            c_desc.border = thin_border

            # Cell C: Step
            c_step = ws.cell(row=current_row, column=3)
            c_step.font = body_font
            c_step.alignment = align_top_left
            c_step.fill = fill
            c_step.border = thin_border

            # Cell D: Expected Result
            c_exp = ws.cell(row=current_row, column=4)
            c_exp.font = body_font
            c_exp.alignment = align_top_left
            c_exp.fill = fill
            c_exp.border = thin_border

            # Cell E: Actual Result (Empty for input)
            c_act = ws.cell(row=current_row, column=5)
            c_act.font = body_font
            c_act.alignment = align_top_left
            c_act.fill = fill
            c_act.border = thin_border

            # Cell F: Status
            c_stat = ws.cell(row=current_row, column=6)
            c_stat.font = Font(name=font_family, size=10, bold=True, color="5B21B6")
            c_stat.alignment = align_top_center
            c_stat.fill = pending_fill
            c_stat.border = thin_border
            
            current_row += 1

    # ================= SHEET 1: OVERVIEW =================
    ws_ov = wb.create_sheet(title="Dashboard & Overview")
    ws_ov.views.sheetView[0].showGridLines = True
    
    # Title
    ws_ov.merge_cells("A1:D1")
    ws_ov["A1"] = "Leave Management System (LMS) - Testing Suite"
    ws_ov["A1"].font = title_font
    ws_ov["A1"].fill = title_fill
    ws_ov["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_ov.row_dimensions[1].height = 45

    ws_ov.append([])
    
    # Metadata Block
    ws_ov.cell(row=3, column=1, value="PROJECT NAME:").font = meta_label_font
    ws_ov.cell(row=3, column=2, value="Leave Management System (LMS)").font = meta_val_font
    ws_ov.cell(row=4, column=1, value="FRAMEWORK:").font = meta_label_font
    ws_ov.cell(row=4, column=2, value="Next.js (App Router) with Prisma & Supabase").font = meta_val_font
    ws_ov.cell(row=5, column=1, value="GENERATION DATE:").font = meta_label_font
    ws_ov.cell(row=5, column=2, value="June 15, 2026").font = meta_val_font
    ws_ov.cell(row=6, column=1, value="TEST TYPE:").font = meta_label_font
    ws_ov.cell(row=6, column=2, value="Manual Functional & Policy Validation Test Suite").font = meta_val_font
    
    # Table of Contents & Statistics
    ws_ov.cell(row=8, column=1, value="Test Suite Index & Statistics").font = Font(name=font_family, size=14, bold=True, color="1E3A8A")
    ws_ov.row_dimensions[8].height = 25
    
    stats_headers = ["Module Tab", "Test ID Range", "Description / Focus", "Test Cases Count"]
    ws_ov.append([]) # spacing
    stats_row_num = 10
    ws_ov.append(stats_headers)
    ws_ov.row_dimensions[stats_row_num].height = 24
    for c_idx in range(1, 5):
        cell = ws_ov.cell(row=stats_row_num, column=c_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    modules_stats = [
        ("Auth & Portal", "LMS-AUT-01 to LMS-POR-03", "User authentication, profile adjustments, employee dashboard, and wellness scores", 8),
        ("Leave Application", "LMS-LVE-01 to LMS-LVE-10", "Applying leaves, half days, sandwich rules, SL document links, probation end date tracking, future projections", 10),
        ("Negative Balances", "LMS-NEG-01 to LMS-NEG-04", "Testing balances falling below zero, maximum negative policies, and block validation", 4),
        ("Comp-Off Management", "LMS-CMP-01 to LMS-CMP-03", "Logging weekend/holiday work hours, approval flows, expired credits, and usage rules", 3),
        ("HR Directory & Ledger", "LMS-DIR-01 to LMS-LED-02", "Employee profiles, LWD notice periods, deletion, CSV sync, leave register, ledger database-first updates", 7),
        ("Holidays & Settings", "LMS-HOL-01 to LMS-SET-03", "Holidays list/calendar views, policy parameters, Year-End Closure reset, system date override test mode", 7),
        ("Reports & Auditing", "LMS-REP-01 to LMS-REP-02", "Leave Register approval audits, database transaction audit logs, CSV exports", 2)
    ]
    
    curr_row = 11
    total_cases = 0
    for mod in modules_stats:
        ws_ov.append([mod[0], mod[1], mod[2], mod[3]])
        ws_ov.row_dimensions[curr_row].height = 22
        total_cases += mod[3]
        for col_idx in range(1, 5):
            cell = ws_ov.cell(row=curr_row, column=col_idx)
            cell.font = body_font
            cell.border = thin_border
            cell.fill = zebra_fill if curr_row % 2 == 0 else white_fill
            if col_idx == 4:
                cell.alignment = align_center
                cell.font = Font(name=font_family, size=10, bold=True)
            else:
                cell.alignment = align_left
        curr_row += 1
        
    # Total row
    ws_ov.append(["TOTAL TEST CASES", "", "", total_cases])
    ws_ov.row_dimensions[curr_row].height = 25
    ws_ov.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=3)
    t_cell_lbl = ws_ov.cell(row=curr_row, column=1)
    t_cell_lbl.font = Font(name=font_family, size=11, bold=True)
    t_cell_lbl.alignment = Alignment(horizontal="right", vertical="center")
    t_cell_lbl.fill = sub_header_fill
    t_cell_lbl.border = thin_border
    
    t_cell_val = ws_ov.cell(row=curr_row, column=4)
    t_cell_val.font = Font(name=font_family, size=11, bold=True, color="1E3A8A")
    t_cell_val.alignment = align_center
    t_cell_val.fill = sub_header_fill
    t_cell_val.border = thin_border
    
    # Format Overview column widths
    ws_ov.column_dimensions["A"].width = 25
    ws_ov.column_dimensions["B"].width = 28
    ws_ov.column_dimensions["C"].width = 75
    ws_ov.column_dimensions["D"].width = 18

    # Execution Instructions Box
    instr_row = curr_row + 3
    ws_ov.cell(row=instr_row, column=1, value="Instructions for Execution:").font = Font(name=font_family, size=12, bold=True, color="1E3A8A")
    ws_ov.cell(row=instr_row+1, column=1, value="1. Use the tabs at the bottom to navigate through the modules.").font = body_font
    ws_ov.cell(row=instr_row+2, column=1, value="2. Execute each test case step-by-step in the system.").font = body_font
    ws_ov.cell(row=instr_row+3, column=1, value="3. Record findings in the 'Actual Result' column and set the 'Status' column (e.g. Pass, Fail, Blocked).").font = body_font
    ws_ov.cell(row=instr_row+4, column=1, value="4. Use the Status fill codes: Pass (Green), Fail (Red), Pending (Amber).").font = body_font

    # ================= SHEET 2: AUTH & PORTAL =================
    ws_ap = wb.create_sheet(title="Auth & Portal")
    format_sheet_common(ws_ap)
    write_header(ws_ap, "Module: User Authentication & Portal Dashboard")
    
    ap_tcs = [
        {
            "id": "LMS-AUT-01",
            "description": "Log in with valid Employee credentials",
            "step": "1. Navigate to /login.\n2. Enter a valid employee email (e.g., employee@example.com) and password.\n3. Click 'Login' or press Enter.",
            "expected": "1. Authentication succeeds.\n2. User session is created.\n3. User is automatically redirected to '/portal' (My Portal).\n4. A success message is shown."
        },
        {
            "id": "LMS-AUT-02",
            "description": "Log in with valid HR/Admin/Manager credentials",
            "step": "1. Navigate to /login.\n2. Enter a valid admin/manager email (e.g., hr@example.com) and password.\n3. Click 'Login'.",
            "expected": "1. Authentication succeeds.\n2. User session is created.\n3. User is automatically redirected to '/' (Admin Dashboard) showing company metrics.\n4. Admin elements are visible."
        },
        {
            "id": "LMS-AUT-03",
            "description": "Log in with invalid credentials",
            "step": "1. Navigate to /login.\n2. Enter a wrong email or invalid password.\n3. Click 'Login'.",
            "expected": "1. Authentication fails.\n2. Toast error displays 'Invalid credentials'.\n3. User remains on /login page; no redirect occurs."
        },
        {
            "id": "LMS-AUT-04",
            "description": "HR/Admin account profile decoupling",
            "step": "1. Log in as HR/Admin.\n2. Verify if 'My Portal' link is visible in the sidebar or navigation header.\n3. Open Leave Ledger page selection dropdown and search for HR users.\n4. Navigate to Team Directory and check if HR user profile is listed.",
            "expected": "1. 'My Portal' link is hidden for Admins (separating personal and corporate roles).\n2. HR user names and profiles do not appear in the employee list, dropdowns, or directory (Rule 3.0)."
        },
        {
            "id": "LMS-AUT-05",
            "description": "Logout functionality and private route guard",
            "step": "1. Log in as any employee or admin.\n2. Click the 'Logout' button in the header.\n3. After logout, attempt to navigate directly to '/portal' or '/team' using the browser address bar.",
            "expected": "1. User is successfully logged out.\n2. Redirected back to '/login'.\n3. Direct access to protected URLs is blocked and redirects to '/login'."
        },
        {
            "id": "LMS-POR-01",
            "description": "Update profile communication email and password",
            "step": "1. Log in as an Employee.\n2. Click on Profile from the navigation menu.\n3. Update the 'Communication Email' to a new address.\n4. Fill in the 'Change Password' fields (Current Password, New Password, Confirm New Password) and submit.",
            "expected": "1. Form submits successfully, displaying a success toast.\n2. Communication email updates in the database (used for notifications).\n3. Password is changed; the user must log in with the new password on next session."
        },
        {
            "id": "LMS-POR-02",
            "description": "Employee Leave Balance Overview validation",
            "step": "1. Log in as Employee and view 'My Portal' dashboard.\n2. Locate the 'Leave Balance Overview' card.\n3. Compare displayed balances (PL, CL, SL, COMP) with those in the database (Prisma leave_balances).",
            "expected": "1. Card renders with four sections for PL, CL, SL, and COMP.\n2. Balances are correct and exactly reflect current DB values.\n3. Layout is responsive and numbers are clean."
        },
        {
            "id": "LMS-POR-03",
            "description": "Wellness Score indicator validation",
            "step": "1. Log in as an Employee.\n2. Inspect the 'Wellness Score' card containing a progress bar.\n3. Verify formula: score is calculated dynamically based on total used leaves vs total allowed entitlement.",
            "expected": "1. Progress bar loads correctly.\n2. If score > 80, shows 'Great job! You\'re balancing work and rest well.'\n3. If score <= 80, shows 'Consider taking some time off soon.' to encourage work-life balance."
        }
    ]
    write_test_cases(ws_ap, ap_tcs)

    # ================= SHEET 3: LEAVE APPLICATION =================
    ws_la = wb.create_sheet(title="Leave Application")
    format_sheet_common(ws_la)
    write_header(ws_la, "Module: Applying for Leave & Policy Validation Rules")
    
    la_tcs = [
        {
            "id": "LMS-LVE-01",
            "description": "Apply for normal Privilege Leave (PL) - working days only",
            "step": "1. Log in as Employee and navigate to 'Apply for Leave' on My Portal.\n2. Select Leave Type: Privilege Leave (PL).\n3. Choose Start Date: Monday, End Date: Wednesday of the same week (3 working days).\n4. Fill in Reason (mandatory).\n5. Click 'Submit Request'.",
            "expected": "1. 'Requested Days' calculates dynamically as 3.0 days.\n2. Request submits successfully, showing success toast.\n3. Request appears in 'Recent Requests' table with status PENDING."
        },
        {
            "id": "LMS-LVE-02",
            "description": "Apply for PL overlapping weekends and public holidays (Exclusion rules)",
            "step": "1. Identify a week containing a public holiday (e.g. Wednesday, June 17, 2026).\n2. Select PL.\n3. Set Start Date: Tuesday (June 16), End Date: Thursday (June 18).\n4. Observe the 'Requested Days' count.",
            "expected": "1. Projected/Requested Days shows 2.0 days, excluding Wednesday (public holiday) and weekends.\n2. Re-verifies that only true working days are debited from the employee's PL balance."
        },
        {
            "id": "LMS-LVE-03",
            "description": "Sick Leave (SL) Medical Certificate Warning triggering",
            "step": "1. Select Leave Type: Sick Leave (SL).\n2. Set dates that span 3 or more working days (e.g., Monday to Wednesday).\n3. Observe the form alerts.",
            "expected": "1. Form displays a warning: 'Medical Certificate Required. Sick leaves exceeding 2 days require a medical certificate. Please paste the link above.'\n2. Highlighted warning box appears dynamically beneath the document link input."
        },
        {
            "id": "LMS-LVE-04",
            "description": "Submit SL exceeding 2 days with medical document URL",
            "step": "1. Select SL for 3 days.\n2. In the 'Web Link / Document URL' input, enter a valid URL (e.g., https://drive.google.com/cert.pdf).\n3. Enter reason and click 'Submit Request'.",
            "expected": "1. Submission succeeds.\n2. Document URL is saved in the database under attachmentUrl field.\n3. Request in 'Recent Requests' displays a clickable 'View Document' link opening the provided URL."
        },
        {
            "id": "LMS-LVE-05",
            "description": "Probation Period PL restriction (Rule 13 / 21)",
            "step": "1. Select or edit an employee such that today's date is before their 'Probation End Date' (or they are within 6 months of joining if no date specified).\n2. Log in as this employee.\n3. Navigate to Apply for Leave and select Privilege Leave (PL).\n4. Try to submit a request.",
            "expected": "1. Form validation prevents submission, or displays warning that PL cannot be requested during the probation period.\n2. Ensures probationers are blocked from utilizing Privilege Leave."
        },
        {
            "id": "LMS-LVE-06",
            "description": "Apply for Casual Leave (CL) exceeding 4 calendar days (Continuous break rule)",
            "step": "1. Select Leave Type: Casual Leave (CL).\n2. Set Start Date: Friday, End Date: next Tuesday (5 calendar days, including weekend).\n3. Check the calculated duration and alerts.",
            "expected": "1. Validation triggers: 'Continuous break exceeds limits. Request automatically converted to PL (Rule 37).'\n2. Leave type on the server projection changes to PL.\n3. Weekends/holidays are NOT counted in the final duration because PL excludes them (final duration is 3 days instead of 5 days CL)."
        },
        {
            "id": "LMS-LVE-07",
            "description": "Apply for CL spanning weekends with Sandwich Rule Enabled",
            "step": "1. In settings, verify the 'weekend_sandwich_rule' is set to 'true'.\n2. Log in as employee. Select Casual Leave (CL).\n3. Set Start Date: Friday, End Date: next Monday (4 calendar days, including Sat & Sun).\n4. Check the calculated days and warning.",
            "expected": "1. Sandwich warning shows: 'Sandwich rule applies: weekends/holidays are included in your requested days.'\n2. Requested days is calculated as 4.0 days (including Saturday and Sunday).\n3. Balance debit will count all 4 calendar days."
        },
        {
            "id": "LMS-LVE-08",
            "description": "Apply for CL spanning weekends with Sandwich Rule Disabled",
            "step": "1. In settings, set the 'weekend_sandwich_rule' to 'false'.\n2. Select Casual Leave (CL).\n3. Set Start Date: Friday, End Date: next Monday (4 calendar days).\n4. Observe calculated days.",
            "expected": "1. Requested days is calculated as 2.0 days (Saturday and Sunday are excluded).\n2. No sandwich warning is displayed. Only working days count."
        },
        {
            "id": "LMS-LVE-09",
            "description": "Apply for Half-Day Leave",
            "step": "1. Select any Leave Type (e.g. PL).\n2. Set 'Half Day' dropdown to 'First Half' or 'Second Half'.\n3. Select Start Date and End Date as the same day.\n4. Click 'Submit Request'.",
            "expected": "1. 'Requested Days' calculates as 0.5 days.\n2. Submit succeeds. Leave request in Recent Requests lists duration as '0.5' and half-day status."
        },
        {
            "id": "LMS-LVE-10",
            "description": "Future Projection Balance calculation",
            "step": "1. Select PL.\n2. Set Start Date to a future month (e.g., 4 months from today).\n3. Observe the projected balance message and calculation.",
            "expected": "1. Warning shows 'Showing projected balance for [Future Date]'.\n2. Projected balance dynamically increases based on month difference * monthly PL accrual rate config (e.g. +6.0 PL for 4 months if rate is 1.5)."
        }
    ]
    write_test_cases(ws_la, la_tcs)

    # ================= SHEET 4: NEGATIVE BALANCES =================
    ws_nb = wb.create_sheet(title="Negative Balances")
    format_sheet_common(ws_nb)
    write_header(ws_nb, "Module: Leave Limits & Negative Balance Rules")
    
    nb_tcs = [
        {
            "id": "LMS-NEG-01",
            "description": "Apply for Leave exceeding active balance (Negative Allowed)",
            "step": "1. Log in as employee with PL balance = 1.0.\n2. In the leave form, select PL and apply for 3 working days.\n3. Check the projected net balance in the form status box.",
            "expected": "1. Projected net balance shows as -2.0.\n2. Dynamic warning is displayed: 'Leave allowed in negative balance. Recovery of 2.0 days may apply upon exit.'"
        },
        {
            "id": "LMS-NEG-02",
            "description": "Submit leave request with negative balance (Under policy limit)",
            "step": "1. From LMS-NEG-01, click the 'Submit (Negative Balance)' button.\n2. Confirm submission.",
            "expected": "1. System allows the request to be submitted.\n2. A warning toast is shown reminding that balance will go negative.\n3. The request is created in DB with isNegative=true and negativeAmount=2.0."
        },
        {
            "id": "LMS-NEG-03",
            "description": "Apply for leave exceeding configured negative limit (Blocked)",
            "step": "1. Ensure MAX_NEGATIVE_LEAVE system config is set to -5.0.\n2. Log in as employee with PL balance = 0.0.\n3. In the leave form, apply for PL for 6 working days (net balance would be -6.0, exceeding -5.0).\n4. Inspect the form controls.",
            "expected": "1. Form shows error: 'Exceeds minimum negative limit (-5 days). Request cannot be submitted.'\n2. Submit button is disabled.\n3. Submitting the form triggers a toast block: 'Cannot apply: net balance would be -6, below the minimum allowed limit of -5 days.'"
        },
        {
            "id": "LMS-NEG-04",
            "description": "Compensatory Off (COMP) negative balance block validation",
            "step": "1. Log in as employee with COMP balance = 0.0.\n2. Select Leave Type: Compensatory Off (COMP).\n3. Apply for 1.0 day of COMP leave.\n4. Attempt to submit.",
            "expected": "1. Form displays that COMP cannot go negative.\n2. Click submit: Toast error shows 'Compensatory Off cannot go into negative balance. Apply for Loss of Pay instead.'\n3. Action is blocked."
        }
    ]
    write_test_cases(ws_nb, nb_tcs)

    # ================= SHEET 5: COMP-OFF MANAGEMENT =================
    ws_co = wb.create_sheet(title="Comp-Off Management")
    format_sheet_common(ws_co)
    write_header(ws_co, "Module: Compensatory Off Work Logs & Approvals")
    
    co_tcs = [
        {
            "id": "LMS-CMP-01",
            "description": "Request Comp-Off credit (Comp-Off Work Log)",
            "step": "1. Log in as Employee and go to 'My Portal'.\n2. In the 'Comp-Off Work Log' card, select Date Worked (e.g. a Saturday).\n3. Input Hours Worked: 8.0 hours.\n4. Input Reason describing weekend work done.\n5. Click 'Submit Log'.",
            "expected": "1. Log submits successfully.\n2. A pending CompOffWorkEntry record is added to the database.\n3. Renders as PENDING in employee recent history log."
        },
        {
            "id": "LMS-CMP-02",
            "description": "Approve Comp-Off Work Log by Admin/Manager",
            "step": "1. Log in as Admin/HR.\n2. Go to requests list or approval dashboard.\n3. Locate the employee's pending Comp-Off request.\n4. Click 'Approve'.",
            "expected": "1. CompOffWorkEntry status updates to APPROVED.\n2. System sets approvedById and approvedAt.\n3. Calculates daysCredited = 1.0 (for 8 hours worked) and adds it to the employee's leave_balances.comp.\n4. Sets an expiryDate (e.g. +60 or +90 days from the work date)."
        },
        {
            "id": "LMS-CMP-03",
            "description": "Expired Comp-Off credit handling",
            "step": "1. Simulate an expired Comp-Off by going to the database and editing the expiryDate of an approved Comp-Off entry to a past date.\n2. Force a ledger rebuild or trigger Year-End Closure / accrual run.",
            "expected": "1. The system marks the Comp-Off as EXPIRED in CompOffWorkEntry.\n2. The active COMP balance of the employee is debited by the expired amount."
        }
    ]
    write_test_cases(ws_co, co_tcs)

    # ================= SHEET 6: HR DIRECTORY & LEDGER =================
    ws_dl = wb.create_sheet(title="HR Team & Ledger")
    format_sheet_common(ws_dl)
    write_header(ws_dl, "Module: HR Team Directory, Ledger & Adjustments")
    
    dl_tcs = [
        {
            "id": "LMS-DIR-01",
            "description": "View Team Directory, search, and CSV export",
            "step": "1. Log in as HR/Admin.\n2. Go to 'Team Directory' (/team).\n3. Search for an employee by name in the search field.\n4. Click 'Export CSV'.",
            "expected": "1. Table displays all employee names, roles, join date, status, probation ends, and current balances.\n2. Searching filters the table in real-time.\n3. Clicking Export CSV downloads a formatted spreadsheet of the employee list."
        },
        {
            "id": "LMS-DIR-02",
            "description": "Add new employee and initial leave balance setup",
            "step": "1. In Team Directory, click 'Add Employee'.\n2. Fill in Name, Email, Department, Role (EMPLOYEE), Status (ACTIVE), Join Date, Probation End Date, and Communication Email.\n3. Click 'Save'.",
            "expected": "1. User profile is successfully created.\n2. A default empty or prorated `leave_balances` row is automatically initialized in the database.\n3. User is sent a confirmation, and appears in the Team Directory grid."
        },
        {
            "id": "LMS-DIR-03",
            "description": "Update Employee Status, Notice Period, and LWD (Last Working Day)",
            "step": "1. In Team Directory, click the 'Edit' icon for an employee.\n2. Change Status to 'NOTICE_PERIOD' or 'RESIGNED'.\n3. Select a 'Last Working Day (LWD)' date from the date picker.\n4. Click 'Save'.",
            "expected": "1. User status is updated successfully in DB (User table: status and lastWorkingDay fields).\n2. Employee remains searchable, showing updated status."
        },
        {
            "id": "LMS-DIR-04",
            "description": "Delete Employee Profile (Cascade check)",
            "step": "1. In Team Directory, click the 'Delete' icon for an employee profile.\n2. Confirm the warning popup.",
            "expected": "1. User is removed or soft-deleted from the directory.\n2. Linked records (leave_balances, requests, ledgerEntries) are cleaned up or cascaded, preventing database foreign key integrity errors."
        },
        {
            "id": "LMS-LED-01",
            "description": "View employee Leave Ledger (Decoupled UUID dropdown)",
            "step": "1. Navigate to Leave Ledger (/ledger) as HR.\n2. Click the employee selection dropdown.\n3. Select a user and examine the ledger grid.",
            "expected": "1. Dropdown shows only human-readable details: Employee Name and Department. Internally, UUIDs are hidden (Rule 5.0).\n2. Grid shows detailed, sequential rows: Opening balance, monthly accruals, approved leaves, manual adjustments.\n3. Performance is fast (< 1s) as it reads from the pre-computed database-first LeaveLedgerEntry table."
        },
        {
            "id": "LMS-LED-02",
            "description": "Update Database ledger sync tool",
            "step": "1. In the Leave Ledger page, click the 'Update Database' button.\n2. Wait for completion.",
            "expected": "1. Triggered API (/api/admin/sync-ledger) recalculates all history.\n2. Success toast appears.\n3. Balances are fully re-synchronized with active leave records."
        },
        {
            "id": "LMS-ADJ-01",
            "description": "Manual Leave Balance Adjustment by Admin",
            "step": "1. Open the Adjustments form (or click edit balance in ledger).\n2. Select Employee, Leave Type: PL.\n3. Select Adjustment Type: AUDIT.\n4. Enter Amount: +2.5 (positive) or -1.0 (negative) and a Reason.\n5. Save the adjustment.",
            "expected": "1. Adjustment is saved in the database.\n2. Active balance immediately updates in `leave_balances`.\n3. An adjustment type entry (e.g. ADJ-PL) is created in `LeaveLedgerEntry` so that it is audited in history."
        }
    ]
    write_test_cases(ws_dl, dl_tcs)

    # ================= SHEET 7: HOLIDAYS & SETTINGS =================
    ws_hs = wb.create_sheet(title="Holidays & Settings")
    format_sheet_common(ws_hs)
    write_header(ws_hs, "Module: Holidays, Policies & Year-End Closure")
    
    hs_tcs = [
        {
            "id": "LMS-HOL-01",
            "description": "Add, Edit, and Delete public holiday (Admin Only)",
            "step": "1. Log in as Admin and go to Holidays (/holidays).\n2. Click 'Add Holiday'. Fill in Name (e.g., Independence Day), Date, Type (Fixed/Floating), and location.\n3. Click Save.\n4. Locate the holiday in the list, edit the name, and save.\n5. Click the delete icon to remove a temporary holiday.",
            "expected": "1. Holidays can be successfully created, edited, and deleted in the Holiday table.\n2. Renders immediately on UI.\n3. Non-admin users are restricted from performing write operations."
        },
        {
            "id": "LMS-HOL-02",
            "description": "Holiday Calendar View and list filters",
            "step": "1. Navigate to Holidays.\n2. Toggle to 'Calendar View' (monthly grid).\n3. Toggle to 'List View' and use the Year Filter dropdown.\n4. Verify CSV Export from Holidays page.",
            "expected": "1. Calendar view displays a monthly grid with holidays highlighted on their dates.\n2. List view filters holidays by year selected.\n3. CSV Export downloads a clean CSV file of all configured public holidays."
        },
        {
            "id": "LMS-SET-01",
            "description": "Update Policy Configurations (Settings)",
            "step": "1. Go to Settings (/settings) -> Policy Config.\n2. Modify variables: PL Accrual Rate (e.g., to 1.8), Accrual Base Days (e.g., 22), Max Carry Forward (e.g., 40), and Negative Leave Limit.\n3. Click 'Save Config'.",
            "expected": "1. Policy adjustments are written to `SystemConfig` table in DB.\n2. Success toast is displayed.\n3. New accrual and validation tasks immediately respect the updated config values."
        },
        {
            "id": "LMS-SET-02",
            "description": "Execute Year-End Closure processing",
            "step": "1. Go to Settings -> Year-End Closure.\n2. Fill in remarks and click 'Run Year-End Reset'.\n3. Review balances after closure.",
            "expected": "1. CL and SL balances are reset back to default opening balances.\n2. PL balances are carried forward up to the config's Max Carry Forward limit; surplus PL is expired.\n3. Creates a record in `LeaveYearClosure` and logs CLOSING/OPENING ledger records."
        },
        {
            "id": "LMS-SET-03",
            "description": "Enable Test Mode and System Date Override (Rule 42)",
            "step": "1. In Settings, locate the 'Test Mode / Date Override' tab.\n2. Toggle Test Mode to ON.\n3. Select a future override date (e.g. December 31, 2026).\n4. Click Save.",
            "expected": "1. System Date Override is saved in DB.\n2. Header indicates system is running in Test Mode.\n3. The application's server tasks evaluate dates relative to the override date (allowing time-travel checks on accruals and closures)."
        }
    ]
    write_test_cases(ws_hs, hs_tcs)

    # ================= SHEET 8: REPORTS & AUDITING =================
    ws_ra = wb.create_sheet(title="Reports & Auditing")
    format_sheet_common(ws_ra)
    write_header(ws_ra, "Module: Reports, Leave Register & Audit Logs")
    
    ra_tcs = [
        {
            "id": "LMS-REP-01",
            "description": "Leave Register Screen review & Approver Audit",
            "step": "1. Log in as HR/Admin.\n2. Navigate to Reports or Leave Register page.\n3. Check the list of leave applications.\n4. Find an approved leave request.",
            "expected": "1. The Leave Register lists all leave applications, statuses, dates, reasons, and employee details.\n2. Approved requests explicitly show the name of the Admin who approved it and the approval timestamp (Rule 1.1 / 1.3 / 7.0).\n3. Re-verifies ledger entry links to the active approver ID."
        },
        {
            "id": "LMS-REP-02",
            "description": "Verify General Audit Logs",
            "step": "1. Navigate to Audit Logs page (/audit).\n2. View the list of system events.\n3. Perform an action (e.g., change a setting or apply for leave) and refresh the audit log.",
            "expected": "1. Audit Log screen displays action, entity name, entity ID, old value, new value, date, and user who triggered it.\n2. The newly performed action appears as the latest entry, confirming auditing is functional."
        }
    ]
    write_test_cases(ws_ra, ra_tcs)

    # Save Workbook
    filename = "LMS_Manual_Test_Cases.xlsx"
    wb.save(filename)
    print(f"Workbook saved successfully as {filename}")

if __name__ == "__main__":
    create_manual_test_cases_excel()
